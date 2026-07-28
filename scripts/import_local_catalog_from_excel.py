import io
import json
import sqlite3
from pathlib import Path

import openpyxl
from PIL import Image as PILImage


WORKBOOK_PATH = Path(
    r"D:/HealthCare/副本医疗设备耗材需求清单（肖利军确定能提供的耗材和少量设备-含图片）-0624(1).xlsx"
)
DATABASE_PATH = Path(r"C:/Users/Michael/Documents/Healthcare/data/healthcare.db")
IMAGE_DIRECTORY = Path(
    r"C:/Users/Michael/Documents/Healthcare/public/uploads/images/excel-batch"
)
CONSUMABLE_LOCALIZATION_PATH = (
    Path(__file__).resolve().parents[1]
    / "data"
    / "consumable-english-localizations.json"
)
DEVICE_LOCALIZATION_PATH = (
    Path(__file__).resolve().parents[1]
    / "data"
    / "device-english-localizations.json"
)
ENGLISH_LOCALIZATIONS = json.loads(
    CONSUMABLE_LOCALIZATION_PATH.read_text(encoding="utf-8")
)
ENGLISH_LOCALIZATIONS.update(
    json.loads(DEVICE_LOCALIZATION_PATH.read_text(encoding="utf-8"))
)


NEW_CATEGORIES = [
    {
        "slug": "protective-consumables",
        "nameZh": "防护类",
        "nameEn": "Protective Consumables",
        "descriptionZh": "医用口罩、手套、防护服、护目镜、面屏、鞋套与帽类产品。",
        "descriptionEn": "Masks, gloves, protective apparel, goggles, face shields, shoe covers, and caps.",
        "sortOrder": 1,
    },
    {
        "slug": "injection-infusion-consumables",
        "nameZh": "注射输液类",
        "nameEn": "Injection & Infusion Consumables",
        "descriptionZh": "注射器、针头、输液器、留置针及相关输注耗材。",
        "descriptionEn": "Syringes, needles, infusion sets, cannulas, and related infusion consumables.",
        "sortOrder": 2,
    },
    {
        "slug": "wound-care-consumables",
        "nameZh": "伤口护理类",
        "nameEn": "Wound Care Consumables",
        "descriptionZh": "敷料、绷带、纱布、棉球、创可贴及基础处理工具。",
        "descriptionEn": "Dressings, bandages, gauze, cotton, adhesive bandages, and basic wound-care tools.",
        "sortOrder": 3,
    },
    {
        "slug": "disinfection-cleaning-consumables",
        "nameZh": "消毒清洁类",
        "nameEn": "Disinfection & Cleaning",
        "descriptionZh": "酒精、碘伏、消毒湿巾、洗手液与消毒凝胶等产品。",
        "descriptionEn": "Alcohol, povidone-iodine, disinfecting wipes, hand sanitizer, and disinfection gels.",
        "sortOrder": 4,
    },
    {
        "slug": "testing-sampling-consumables",
        "nameZh": "检验采样类",
        "nameEn": "Testing & Sampling Consumables",
        "descriptionZh": "采血管、快速检测试剂与相关采样耗材。",
        "descriptionEn": "Blood collection tubes, rapid test kits, and related sampling consumables.",
        "sortOrder": 5,
    },
    {
        "slug": "basic-care-consumables",
        "nameZh": "基础护理类",
        "nameEn": "Basic Care Consumables",
        "descriptionZh": "导管、尿袋、基础护理包材及病房护理相关耗材。",
        "descriptionEn": "Catheters, urine bags, nursing disposables, and ward-care consumables.",
        "sortOrder": 6,
    },
    {
        "slug": "monitoring-diagnostic-devices",
        "nameZh": "监护诊断设备",
        "nameEn": "Monitoring & Diagnostic Devices",
        "descriptionZh": "监护仪、心电、超声及其他基础诊断设备。",
        "descriptionEn": "Patient monitors, ECG systems, ultrasound, and other diagnostic devices.",
        "sortOrder": 7,
    },
    {
        "slug": "treatment-infusion-devices",
        "nameZh": "治疗输注设备",
        "nameEn": "Treatment & Infusion Devices",
        "descriptionZh": "治疗输注、药液控制及相关治疗设备。",
        "descriptionEn": "Treatment, infusion control, and related therapeutic devices.",
        "sortOrder": 8,
    },
    {
        "slug": "respiratory-emergency-devices",
        "nameZh": "呼吸急救设备",
        "nameEn": "Respiratory & Emergency Devices",
        "descriptionZh": "呼吸支持、除颤、制氧与急救吸引设备。",
        "descriptionEn": "Respiratory support, defibrillation, oxygen, and emergency suction devices.",
        "sortOrder": 9,
    },
    {
        "slug": "anesthesia-surgical-devices",
        "nameZh": "手术麻醉设备",
        "nameEn": "Anesthesia Devices",
        "descriptionZh": "麻醉机与相关手术麻醉设备。",
        "descriptionEn": "Anesthesia machines and related perioperative anesthesia devices.",
        "sortOrder": 10,
    },
    {
        "slug": "minimally-invasive-surgical-devices",
        "nameZh": "微创外科设备",
        "nameEn": "Minimally Invasive Surgical Devices",
        "descriptionZh": "超声刀、吻合器及相关微创外科系统与配套部件。",
        "descriptionEn": "Ultrasonic surgery systems, staplers, and related minimally invasive surgical platforms.",
        "sortOrder": 11,
    },
    {
        "slug": "surgical-visualization-devices",
        "nameZh": "手术可视化设备",
        "nameEn": "Surgical Visualization Devices",
        "descriptionZh": "内窥镜影像、手术显微镜与术中可视化设备。",
        "descriptionEn": "Endoscopy imaging, surgical microscopes, and intraoperative visualization devices.",
        "sortOrder": 12,
    },
    {
        "slug": "robotics-navigation-devices",
        "nameZh": "手术机器人与导航设备",
        "nameEn": "Robotics & Navigation Devices",
        "descriptionZh": "手术机器人、导航定位与介入控制设备。",
        "descriptionEn": "Surgical robots, navigation systems, and interventional control devices.",
        "sortOrder": 13,
    },
    {
        "slug": "specialty-treatment-devices",
        "nameZh": "专科诊疗设备",
        "nameEn": "Specialty Treatment Devices",
        "descriptionZh": "眼科等专科治疗设备。",
        "descriptionEn": "Specialty treatment equipment including ophthalmic platforms.",
        "sortOrder": 14,
    },
    {
        "slug": "ward-care-devices",
        "nameZh": "病房护理设备",
        "nameEn": "Ward Care Devices",
        "descriptionZh": "病床等病房护理类设备。",
        "descriptionEn": "Ward-care devices such as hospital beds.",
        "sortOrder": 15,
    },
]


WEBSITE_SLUG_CATEGORY_MAP = {
    "1-1-y16": "minimally-invasive-surgical-devices",
    "1-2-hp401": "minimally-invasive-surgical-devices",
    "1-3-hp501": "minimally-invasive-surgical-devices",
    "1-4-sg": "minimally-invasive-surgical-devices",
    "1-5-ss": "minimally-invasive-surgical-devices",
    "1-6-ocbsgbl": "minimally-invasive-surgical-devices",
    "1-7-hifcbsgpl22": "minimally-invasive-surgical-devices",
    "1-8-hifcbsgpl35": "minimally-invasive-surgical-devices",
    "1-9-hifcbsgpl45": "minimally-invasive-surgical-devices",
    "1-10-ifcbsgpl22": "minimally-invasive-surgical-devices",
    "1-11-ifcbsgpl35": "minimally-invasive-surgical-devices",
    "1-12-ifcbsgpl45": "minimally-invasive-surgical-devices",
    "2-1": "minimally-invasive-surgical-devices",
    "2-2": "minimally-invasive-surgical-devices",
    "2-3": "minimally-invasive-surgical-devices",
    "2-4": "minimally-invasive-surgical-devices",
    "3-1-hypixel-ux5-4k": "surgical-visualization-devices",
    "4-1-oms3500": "surgical-visualization-devices",
    "4-2-oms2350": "surgical-visualization-devices",
    "5-1-rosewood": "specialty-treatment-devices",
    "6-1-consona-n9": "monitoring-diagnostic-devices",
    "6-2-resona-i9": "monitoring-diagnostic-devices",
    "7-1-toumai": "robotics-navigation-devices",
    "7-2-skywalker": "robotics-navigation-devices",
    "7-3-r-one": "robotics-navigation-devices",
    "8-1-excelim-116": "robotics-navigation-devices",
    "8-2-neuronav-118": "robotics-navigation-devices",
}


PRODUCT_NAME_CATEGORY_MAP = {
    "医用外科手套": "protective-consumables",
    "医用口罩及个人防护装备": "protective-consumables",
    "医用检查手套": "protective-consumables",
    "医用外科口罩": "protective-consumables",
    "医用防护服（隔离衣）": "protective-consumables",
    "医用鞋套": "protective-consumables",
    "医用一次性帽子": "protective-consumables",
    "医用护目镜": "protective-consumables",
    "医用防护面屏": "protective-consumables",
    "注射器及针头": "injection-infusion-consumables",
    "静脉输液器及留置针": "injection-infusion-consumables",
    "一次性输液器": "injection-infusion-consumables",
    "采血针": "injection-infusion-consumables",
    "伤口敷料及绷带": "wound-care-consumables",
    "医用止血带": "wound-care-consumables",
    "医用胶带": "wound-care-consumables",
    "创可贴": "wound-care-consumables",
    "医用纱布": "wound-care-consumables",
    "医用棉球": "wound-care-consumables",
    "医用剪刀": "wound-care-consumables",
    "医用镊子": "wound-care-consumables",
    "医用酒精": "disinfection-cleaning-consumables",
    "碘伏": "disinfection-cleaning-consumables",
    "消毒湿巾": "disinfection-cleaning-consumables",
    "免洗洗手液": "disinfection-cleaning-consumables",
    "医用消毒凝胶": "disinfection-cleaning-consumables",
    "真空采血管": "testing-sampling-consumables",
    "快速诊断检测试剂盒": "testing-sampling-consumables",
    "医用尿袋": "basic-care-consumables",
    "医用导管": "basic-care-consumables",
    "医用手术洞巾及手术衣": "basic-care-consumables",
    "医用体温计": "basic-care-consumables",
    "医用血压计": "basic-care-consumables",
    "医用听诊器": "basic-care-consumables",
    "医用床单": "basic-care-consumables",
    "医用枕套": "basic-care-consumables",
    "医用垃圾袋": "basic-care-consumables",
    "医用锐器盒": "basic-care-consumables",
    "病人监护仪": "monitoring-diagnostic-devices",
    "心电图机": "monitoring-diagnostic-devices",
    "超声系统": "monitoring-diagnostic-devices",
    "输液泵": "treatment-infusion-devices",
    "呼吸机": "respiratory-emergency-devices",
    "除颤仪": "respiratory-emergency-devices",
    "制氧机": "respiratory-emergency-devices",
    "吸引器": "respiratory-emergency-devices",
    "麻醉机": "anesthesia-surgical-devices",
    "病床": "ward-care-devices",
}


def upsert_categories(connection: sqlite3.Connection):
    category_ids = {}
    for category in NEW_CATEGORIES:
        connection.execute(
            """
            INSERT INTO categories (slug, name_zh, name_en, description_zh, description_en, sort_order)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(slug) DO UPDATE SET
              name_zh = excluded.name_zh,
              name_en = excluded.name_en,
              description_zh = excluded.description_zh,
              description_en = excluded.description_en,
              sort_order = excluded.sort_order
            """,
            (
                category["slug"],
                category["nameZh"],
                category["nameEn"],
                category["descriptionZh"],
                category["descriptionEn"],
                category["sortOrder"],
            ),
        )
        row = connection.execute(
            "SELECT id FROM categories WHERE slug = ?", (category["slug"],)
        ).fetchone()
        category_ids[category["slug"]] = row[0]
    return category_ids


def extract_sheet_images(worksheet, sheet_prefix: str):
    IMAGE_DIRECTORY.mkdir(parents=True, exist_ok=True)
    image_map = {}

    for image in getattr(worksheet, "_images", []):
        anchor = getattr(image, "anchor", None)
        marker = getattr(anchor, "_from", None)
        if marker is None:
            continue

        row_number = marker.row + 1
        image_bytes = image._data()
        local_path = IMAGE_DIRECTORY / f"{sheet_prefix}-{row_number}.png"

        with PILImage.open(io.BytesIO(image_bytes)) as original:
            export = original.convert("RGB")
            export.thumbnail((960, 960))
            canvas = PILImage.new("RGB", (960, 960), "white")
            canvas.paste(
                export,
                ((960 - export.width) // 2, (960 - export.height) // 2),
            )
            canvas.save(local_path, format="PNG", optimize=True)

        image_map[row_number] = f"/uploads/images/excel-batch/{local_path.name}"

    return image_map


def build_product_payload(sheet_name: str, row_number: int, row, image_url: str, category_id: int):
    name = (row[0] or "").strip()
    use_scene = (row[1] or "").strip()
    demand_level = (row[2] or "").strip()
    description = (row[3] or "").strip()
    manufacturer = (row[5] or "").strip()
    model = (row[6] or "").strip()
    notes = (row[7] or "").strip()

    slug_prefix = "excel-consumable" if "耗材" in sheet_name else "excel-device"
    slug = f"{slug_prefix}-{row_number:03d}"
    manufacturer_value = manufacturer or "待补充"
    model_value = model or name
    summary = description or use_scene or name
    application = use_scene or "来自 Excel 清单导入"
    specifications_lines = [f"需求等级：{demand_level}"] if demand_level else []
    if manufacturer:
        specifications_lines.append(f"厂家：{manufacturer}")
    if model:
        specifications_lines.append(f"型号：{model}")
    if notes:
        specifications_lines.append(f"补充说明：{notes}")
    specifications = "\n".join(specifications_lines) or "待补充"
    packaging_lines = [f"来源：{WORKBOOK_PATH.name}"]
    if sheet_name:
        packaging_lines.append(f"Sheet：{sheet_name}")
    packaging = "\n".join(packaging_lines)
    english = ENGLISH_LOCALIZATIONS.get(name, {})

    return {
        "slug": slug,
        "categoryId": category_id,
        "manufacturerZh": manufacturer_value,
        "manufacturerEn": english.get("manufacturerEn", manufacturer_value),
        "model": model_value,
        "nameZh": name,
        "nameEn": english.get("nameEn", name),
        "summaryZh": summary,
        "summaryEn": english.get("summaryEn", summary),
        "applicationZh": application,
        "applicationEn": english.get("applicationEn", application),
        "specificationsZh": specifications,
        "specificationsEn": english.get("specificationsEn", specifications),
        "packagingZh": packaging,
        "packagingEn": english.get("packagingEn", packaging),
        "imageUrl": image_url or "",
        "featured": 0,
        "seoTitleZh": name,
        "seoTitleEn": english.get("nameEn", name),
        "seoDescriptionZh": summary,
        "seoDescriptionEn": english.get("summaryEn", summary),
    }


def upsert_product(connection: sqlite3.Connection, payload):
    connection.execute(
        """
        INSERT INTO products (
          category_id, slug, manufacturer_zh, manufacturer_en, model, name_zh, name_en,
          summary_zh, summary_en, application_zh, application_en, specifications_zh,
          specifications_en, packaging_zh, packaging_en, image_url, featured,
          seo_title_zh, seo_title_en, seo_description_zh, seo_description_en
        )
        VALUES (
          :categoryId, :slug, :manufacturerZh, :manufacturerEn, :model, :nameZh, :nameEn,
          :summaryZh, :summaryEn, :applicationZh, :applicationEn, :specificationsZh,
          :specificationsEn, :packagingZh, :packagingEn, :imageUrl, :featured,
          :seoTitleZh, :seoTitleEn, :seoDescriptionZh, :seoDescriptionEn
        )
        ON CONFLICT(slug) DO UPDATE SET
          category_id = excluded.category_id,
          manufacturer_zh = excluded.manufacturer_zh,
          manufacturer_en = excluded.manufacturer_en,
          model = excluded.model,
          name_zh = excluded.name_zh,
          name_en = excluded.name_en,
          summary_zh = excluded.summary_zh,
          summary_en = excluded.summary_en,
          application_zh = excluded.application_zh,
          application_en = excluded.application_en,
          specifications_zh = excluded.specifications_zh,
          specifications_en = excluded.specifications_en,
          packaging_zh = excluded.packaging_zh,
          packaging_en = excluded.packaging_en,
          image_url = excluded.image_url,
          featured = excluded.featured,
          seo_title_zh = excluded.seo_title_zh,
          seo_title_en = excluded.seo_title_en,
          seo_description_zh = excluded.seo_description_zh,
          seo_description_en = excluded.seo_description_en
        """,
        payload,
    )


def reassign_existing_website_products(connection: sqlite3.Connection, category_ids):
    rows = connection.execute(
        "SELECT id, slug, name_zh FROM products WHERE slug NOT LIKE 'excel-%'"
    ).fetchall()
    updated = 0
    for product_id, slug, name_zh in rows:
        target_slug = WEBSITE_SLUG_CATEGORY_MAP.get(slug) or PRODUCT_NAME_CATEGORY_MAP.get(
            name_zh
        )
        if not target_slug:
            continue
        connection.execute(
            "UPDATE products SET category_id = ? WHERE id = ?",
            (category_ids[target_slug], product_id),
        )
        updated += 1
    return updated


def import_excel_products(connection: sqlite3.Connection, category_ids):
    workbook = openpyxl.load_workbook(WORKBOOK_PATH)
    imported = 0
    for worksheet in workbook.worksheets:
        sheet_prefix = "consumable" if "耗材" in worksheet.title else "device"
        image_map = extract_sheet_images(worksheet, sheet_prefix)
        for row_number in range(2, worksheet.max_row + 1):
            row_values = [worksheet.cell(row_number, column).value for column in range(1, 9)]
            row_values = [value.strip() if isinstance(value, str) else value for value in row_values]
            name = row_values[0]
            if not name:
                continue

            category_slug = PRODUCT_NAME_CATEGORY_MAP.get(name)
            if not category_slug:
                raise RuntimeError(f"Missing category mapping for {name}")

            payload = build_product_payload(
                worksheet.title,
                row_number,
                row_values,
                image_map.get(row_number, ""),
                category_ids[category_slug],
            )
            upsert_product(connection, payload)
            imported += 1
    return imported


def cleanup_legacy_categories(connection: sqlite3.Connection):
    keep_slugs = tuple(category["slug"] for category in NEW_CATEGORIES)
    placeholders = ",".join("?" for _ in keep_slugs)
    connection.execute(
        f"DELETE FROM categories WHERE slug NOT IN ({placeholders}) AND id NOT IN (SELECT DISTINCT category_id FROM products)",
        keep_slugs,
    )


def main():
    connection = sqlite3.connect(DATABASE_PATH)
    try:
        connection.execute("BEGIN")
        category_ids = upsert_categories(connection)
        existing_updated = reassign_existing_website_products(connection, category_ids)
        imported = import_excel_products(connection, category_ids)
        cleanup_legacy_categories(connection)
        connection.commit()
        category_count = connection.execute("SELECT COUNT(*) FROM categories").fetchone()[0]
        product_count = connection.execute("SELECT COUNT(*) FROM products").fetchone()[0]
        print(
            {
                "updatedExistingProducts": existing_updated,
                "importedExcelProducts": imported,
                "categoryCount": category_count,
                "productCount": product_count,
            }
        )
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()


if __name__ == "__main__":
    main()
