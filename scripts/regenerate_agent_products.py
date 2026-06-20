from pathlib import Path
from urllib.parse import urlparse
from urllib.request import Request, urlopen
import os
import ssl

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill


ROWS = {
    3: {
        "model": "以诺康 超声软组织切割止血手术设备（第一代5mm / 第二代7mm）+ OCBSGBL系列二代枪式刀头",
        "spec": "参考规格：主机+换能器+手推车；刀头有效长度22/35/45cm；刀杆直径5-7mm；官方公开信息显示第一代可闭合5mm血管、第二代可安全凝闭最大7mm血管。",
        "image_url": "https://www.innolcon.com/pic/202210131665642108290.png",
        "source_url": "https://www.innolcon.com/page/products/\nhttps://www.innolcon.com/page/knife02/",
        "note": "中等置信度。原表写的是“主机+耗材、兼容强生、5/7mm血管闭合”，与以诺康官网外科超声能量产品线高度匹配；但原表未给出单一注册型号，因此这里按“产品系列/组合”补全。",
    },
    4: {
        "model": "威克 一次性腔镜用直线型全电动切割吻合器及钉仓组件；可同时覆盖其电动款/手动款腔镜直线切割吻合器系列",
        "spec": "参考规格：全电动款支持压力实时反馈、自动按键操控、通配型钉仓；手动/电动系列公开资料显示存在三款器械、三个系列、七种不同高度钉仓；ECC超薄组织专用款成型高度1.0mm；支持鹰嘴型钉仓。",
        "image_url": "https://www.victormedic.com/upload/cn/image/2023-08/col13/1691630678972.png",
        "source_url": "https://www.victormedic.com/col13/9\nhttps://www.victormedic.com/col13/1",
        "note": "中等置信度。原表把手动款和电动款放在同一行，因此这里补全为“威克对应的腔镜直线切割吻合器系列”，而不是强行写成单一SKU。",
    },
    5: {
        "model": "迈瑞 HyPixel UX5 4K三维内窥镜荧光摄影系统",
        "spec": "参考规格：4K+3D一体化平台，支持NIR/荧光成像；官方新闻稿公开提到5.6mm大瞳距双目3D成像、双路真4K、主动除雾、白光+ICG荧光一体化、自动旋转回正。",
        "image_url": "https://www.mindray.com/content/dam/xpace/en/products-solutions/products/laparoscopic-products/endoscope-camera-system/ux5/master1.jpg",
        "source_url": "https://www.mindray.com/cn/products/laparoscopic-products/hypixel-ux5\nhttps://www.mindray.com/cn/media-center/news/mindray-releases-a-4k-3d-electronic-thoracoabdominal-endoscope-system",
        "note": "高置信度。原表的“3D高清腹腔镜、迈瑞、白光+ICG、自动回正、64-71万元”与迈瑞 UX5 官方公开描述高度一致。",
    },
    6: {
        "model": "速迈 OMS3500 手术显微镜",
        "spec": "参考规格：电动变倍系统0.4x-2.4x（1:6），总放大倍数1.3x-18.5x；VarioDist电动变焦物镜F=200-600mm；0°-190°变角双目镜筒；最高照度≥100000 lx；可选内置4K或4K-3D影像系统。",
        "image_url": "https://www.zumax.cn/wp-content/uploads/2024/02/1-7.jpg",
        "source_url": "https://www.zumax.cn/product/oms3500%E6%89%8B%E6%9C%AF%E6%98%BE%E5%BE%AE%E9%95%9C",
        "note": "中高置信度。原表仅写“显微镜、苏州速迈、对标蔡司、带4K摄像/荧光/电动变倍”，速迈官网中 OMS3500 与该描述最接近，且覆盖神外、骨科等科室。",
    },
    7: {
        "model": "以诺康 ROSEWOOD® 眼科超声乳化手术系统",
        "spec": "参考规格：超乳主机+稳定流体/超声控制模块+附件系统；官方公开信息强调前房稳定、术中精控、液流稳定、成本可控、操作便捷；2021年获南德欧盟CE认证。",
        "image_url": "https://www.innolcon.com/templates/default/images/icon31.png",
        "source_url": "https://www.innolcon.com/page/ykcpx/\nhttps://www.innolcon.com/news/html/?30.html%2F=",
        "note": "高置信度。原表写的是“苏州以诺康、眼科白内障超乳、兼容通用耗材、六晶片低发热手柄、双传感液流”，与以诺康 ROSEWOOD 官方产品线和获批新闻高度一致。",
    },
    8: {
        "model": "迈瑞 Consona N9（中端代表机型）/ 昆仑 Resona I9（中高端代表机型）",
        "spec": "参考规格：Consona N系列为智慧型彩超，适合通科筛查与日常全科应用；Resona I9 为高端全身机，覆盖腹部、妇产、心脏、血管、介入等场景。原表提到的AI调参、弹性成像、三维四维、全科室适配，更接近迈瑞高配通用彩超产品组合。",
        "image_url": "https://www.mindray.com/content/dam/xpace/en/products-solutions/products/ultrasound/primary-care/consona-n-series/consona-n9-kv-pc.jpg",
        "source_url": "https://www.mindray.com/cn/products/ultrasound/primary-care/consona-n-series\nhttps://www.mindray.com/cn/products/ultrasound/general-imaging/resona-i9",
        "note": "中等置信度。原表只写“高清B超、迈瑞、全科室适配”，且价格跨中端到中高端，因此更稳妥的补全方式是给出一组最可能的代表系列，而不是冒充唯一型号。",
    },
}


def download(url: str, dest: Path) -> None:
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(req, timeout=30, context=ctx) as response, open(dest, "wb") as fh:
        fh.write(response.read())


def main() -> None:
    desktop = Path.home() / "Desktop"
    source_candidates = [
        p
        for p in desktop.glob("*0617*.xlsx")
        if not p.name.startswith("~$")
        and not p.name.startswith("agent_products_")
    ]
    source = source_candidates[0]
    output = desktop / "agent_products_0617_enriched_v2.xlsx"
    img_dir = Path(r"C:\Users\Michael\Documents\Healthcare\_tmp_product_images")
    img_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(source)
    for sheet_name in ["Enriched", "Sources"]:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    original = wb[wb.sheetnames[0]]
    enriched = wb.create_sheet("Enriched")
    for row in original.iter_rows():
        for cell in row:
            enriched[cell.coordinate].value = cell.value

    enriched["A1"] = "代理产品信息完善建议（基于 2026-06-17 厂家官网公开资料整理）"
    extra_headers = [
        "建议具体型号/系列",
        "参考规格/关键参数",
        "图片链接",
        "来源链接",
        "匹配说明",
        "官网图片预览",
    ]
    for idx, title in enumerate(extra_headers, start=8):
        enriched.cell(row=2, column=idx, value=title)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    sub_fill = PatternFill("solid", fgColor="D9EAF7")

    for col in range(1, 14):
        top = enriched.cell(row=1, column=col)
        top.fill = sub_fill
        top.font = Font(bold=True)
        top.alignment = Alignment(vertical="center")

        header = enriched.cell(row=2, column=col)
        header.fill = header_fill
        header.font = header_font
        header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, info in ROWS.items():
        values = [info["model"], info["spec"], info["image_url"], info["source_url"], info["note"]]
        for col_idx, value in enumerate(values, start=8):
            cell = enriched.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        enriched.cell(row=row_idx, column=10).hyperlink = info["image_url"]
        enriched.cell(row=row_idx, column=10).style = "Hyperlink"
        enriched.cell(row=row_idx, column=11).hyperlink = info["source_url"].split("\n")[0]
        enriched.cell(row=row_idx, column=11).style = "Hyperlink"

        ext = os.path.splitext(urlparse(info["image_url"]).path)[1] or ".png"
        image_path = img_dir / f"row{row_idx}{ext}"
        if not image_path.exists():
            download(info["image_url"], image_path)

        image = XLImage(str(image_path))
        max_width = 180
        scale = max_width / image.width
        image.width = max_width
        image.height = int(image.height * scale)
        enriched.add_image(image, f"M{row_idx}")
        enriched.row_dimensions[row_idx].height = max(110, image.height * 0.75)

    widths = {
        "A": 8,
        "B": 22,
        "C": 16,
        "D": 18,
        "E": 18,
        "F": 16,
        "G": 32,
        "H": 34,
        "I": 42,
        "J": 30,
        "K": 32,
        "L": 42,
        "M": 28,
    }
    for col, width in widths.items():
        enriched.column_dimensions[col].width = width

    enriched.freeze_panes = "A3"
    enriched.auto_filter.ref = f"A2:M{enriched.max_row}"

    sources = wb.create_sheet("Sources")
    sources.append(["序号", "设备/耗材名称", "建议型号/系列", "来源链接"])
    for cell in sources[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row_idx in range(3, 9):
        sources.append(
            [
                original.cell(row=row_idx, column=1).value,
                original.cell(row=row_idx, column=2).value,
                ROWS[row_idx]["model"],
                ROWS[row_idx]["source_url"],
            ]
        )

    for col, width in {"A": 8, "B": 22, "C": 40, "D": 80}.items():
        sources.column_dimensions[col].width = width
    for row in sources.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(output)
    print(output)


if __name__ == "__main__":
    main()
