from pathlib import Path

import openpyxl


SRC = Path(r"C:\Users\Michael\Desktop\agent_products_0617_enriched_official_verified.xlsx")
OUT = Path(r"C:\Users\Michael\Desktop\agent_products_0617_enriched_official_verified_links.xlsx")


LINK_UPDATES = {
    "5-1": {
        "图片链接": "https://www.innolcon.com/page/products/",
        "确定性说明": "高。图片链接规范为以诺康官方产品页，该页可直接查看 ROSEWOOD 眼科超声乳化手术系统官方产品图。",
    },
    "8-1": {
        "图片链接": "https://www.digi-medical.com.cn/News/show?id=79&status=1",
        "确定性说明": "中高。图片链接规范为复旦数字医疗官方新闻实拍页，可直接查看 excelim-116 设备展台实拍图。",
    },
    "8-2": {
        "图片链接": "https://www.digi-medical.com.cn/News/show?id=65&status=1",
        "确定性说明": "中高。图片链接规范为复旦数字医疗官方新闻实拍页，可直接查看 NeuroNav-118 临床/设备实拍图。",
    },
}


def main():
    wb = openpyxl.load_workbook(SRC)
    ws = wb["Enriched"]
    ws["A1"] = "代理产品信息完善建议（按具体型号/具体产品逐行拆分，基于 2026-06-18 官方公开资料复核；图片链接已规范）"

    headers = {ws.cell(2, c).value: c for c in range(1, ws.max_column + 1)}

    for r in range(3, ws.max_row + 1):
        seq = ws.cell(r, 1).value
        if seq not in LINK_UPDATES:
            continue
        for header, value in LINK_UPDATES[seq].items():
            ws.cell(r, headers[header]).value = value

    wb.save(OUT)
    print(str(OUT))


if __name__ == "__main__":
    main()
