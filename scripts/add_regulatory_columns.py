from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


REG_INFO = {
    "1-1": {
        "reg_no": "国械注准20203010682",
        "reg_name": "超声软组织切割止血手术设备",
        "class": "III类",
        "holder": "以诺康医疗科技（苏州）有限公司",
        "scope": "用于软组织切割止血，并可闭合直径不超过5mm的血管。",
        "status": "已用 NMPA 公开记录和厂家官网双重核对。",
    },
    "1-2": {
        "reg_no": "配套部件；见主机注册证",
        "reg_name": "随 Y16 / Y20 主机系统注册核验",
        "class": "按配套主机对应 III类",
        "holder": "以诺康医疗科技（苏州）有限公司/股份有限公司",
        "scope": "作为超声刀系统配套换能器使用。",
        "status": "官网能确认型号 HP401，但独立注册证号未见公开披露，通常需看注册证附件或说明书。",
    },
    "1-3": {
        "reg_no": "配套部件；见主机注册证",
        "reg_name": "随 Y16 主机系统注册核验",
        "class": "按配套主机对应 III类",
        "holder": "以诺康医疗科技（苏州）有限公司",
        "scope": "作为剪式超声刀头配套换能器使用。",
        "status": "官网能确认型号 HP501，但独立注册证号未见公开披露，通常需看注册证附件或说明书。",
    },
    "1-4": {
        "reg_no": "国械注准20253011538；相关公开记录另见国械注准20263010431",
        "reg_name": "一次性使用超声软组织手术刀头 / 超声软组织手术刀头",
        "class": "III类",
        "holder": "以诺康医疗科技（苏州）股份有限公司",
        "scope": "用于超声软组织切割止血系统的一次性刀头耗材。",
        "status": "NMPA 已能锁定到刀头产品注册证；SG 子型号需以注册证附件进一步核验。",
    },
    "1-5": {
        "reg_no": "国械注准20253011538；相关公开记录另见国械注准20263010431",
        "reg_name": "一次性使用超声软组织手术刀头 / 超声软组织手术刀头",
        "class": "III类",
        "holder": "以诺康医疗科技（苏州）股份有限公司",
        "scope": "用于超声软组织切割止血系统的一次性刀头耗材。",
        "status": "NMPA 已能锁定到刀头产品注册证；SS 子型号需以注册证附件进一步核验。",
    },
    "1-6": {
        "reg_no": "国械注准20253011538；相关公开记录另见国械注准20263010431",
        "reg_name": "一次性使用超声软组织手术刀头 / 超声软组织手术刀头",
        "class": "III类",
        "holder": "以诺康医疗科技（苏州）股份有限公司",
        "scope": "用于超声软组织切割止血系统的一次性刀头耗材。",
        "status": "NMPA 已能锁定到刀头产品注册证；OCBSGBL 系列需以注册证附件进一步核验。",
    },
    "1-7": {
        "reg_no": "国械注准20253011538；相关公开记录另见国械注准20263010431",
        "reg_name": "一次性使用超声软组织手术刀头 / 超声软组织手术刀头",
        "class": "III类",
        "holder": "以诺康医疗科技（苏州）股份有限公司",
        "scope": "用于超声软组织切割止血系统的一次性刀头耗材。",
        "status": "NMPA 已能锁定到刀头产品注册证；HIFCBSGPL22 需以注册证附件进一步核验。",
    },
    "1-8": {
        "reg_no": "国械注准20253011538；相关公开记录另见国械注准20263010431",
        "reg_name": "一次性使用超声软组织手术刀头 / 超声软组织手术刀头",
        "class": "III类",
        "holder": "以诺康医疗科技（苏州）股份有限公司",
        "scope": "用于超声软组织切割止血系统的一次性刀头耗材。",
        "status": "NMPA 已能锁定到刀头产品注册证；HIFCBSGPL35 需以注册证附件进一步核验。",
    },
    "1-9": {
        "reg_no": "国械注准20253011538；相关公开记录另见国械注准20263010431",
        "reg_name": "一次性使用超声软组织手术刀头 / 超声软组织手术刀头",
        "class": "III类",
        "holder": "以诺康医疗科技（苏州）股份有限公司",
        "scope": "用于超声软组织切割止血系统的一次性刀头耗材。",
        "status": "NMPA 已能锁定到刀头产品注册证；HIFCBSGPL45 需以注册证附件进一步核验。",
    },
    "1-10": {
        "reg_no": "国械注准20253011538；相关公开记录另见国械注准20263010431",
        "reg_name": "一次性使用超声软组织手术刀头 / 超声软组织手术刀头",
        "class": "III类",
        "holder": "以诺康医疗科技（苏州）股份有限公司",
        "scope": "用于超声软组织切割止血系统的一次性刀头耗材。",
        "status": "NMPA 已能锁定到刀头产品注册证；IFCBSGPL22 需以注册证附件进一步核验。",
    },
    "1-11": {
        "reg_no": "国械注准20253011538；相关公开记录另见国械注准20263010431",
        "reg_name": "一次性使用超声软组织手术刀头 / 超声软组织手术刀头",
        "class": "III类",
        "holder": "以诺康医疗科技（苏州）股份有限公司",
        "scope": "用于超声软组织切割止血系统的一次性刀头耗材。",
        "status": "NMPA 已能锁定到刀头产品注册证；IFCBSGPL35 需以注册证附件进一步核验。",
    },
    "1-12": {
        "reg_no": "国械注准20253011538；相关公开记录另见国械注准20263010431",
        "reg_name": "一次性使用超声软组织手术刀头 / 超声软组织手术刀头",
        "class": "III类",
        "holder": "以诺康医疗科技（苏州）股份有限公司",
        "scope": "用于超声软组织切割止血系统的一次性刀头耗材。",
        "status": "NMPA 已能锁定到刀头产品注册证；IFCBSGPL45 需以注册证附件进一步核验。",
    },
    "2-1": {
        "reg_no": "国械注准20253010808（最接近）；另检索到国械注准20253012206",
        "reg_name": "一次性使用电动腔镜直线型切割吻合器及钉仓组件",
        "class": "III类",
        "holder": "常州威克医疗器械有限公司",
        "scope": "用于腔镜下组织离断、切除和缝合的电动吻合器系统。",
        "status": "已在 NMPA 锁定到威克电动腔镜吻合器注册证；“全电动”与具体证号的逐一映射仍建议再看证书附件。",
    },
    "2-2": {
        "reg_no": "国械注准20243011883",
        "reg_name": "一次性使用电动式腔镜直线型切割吻合器及钉仓",
        "class": "III类",
        "holder": "常州威克医疗器械有限公司",
        "scope": "用于腔镜下组织离断、切除和缝合的电动式吻合器系统。",
        "status": "已在 NMPA 精确检索到产品名和企业名。",
    },
    "2-3": {
        "reg_no": "国械注准20243020427",
        "reg_name": "一次性腔镜用直线型切割吻合器及组件",
        "class": "III类",
        "holder": "常州威克医疗器械有限公司",
        "scope": "用于腔镜下组织离断、切除和缝合的手动吻合器系统。",
        "status": "已在 NMPA 精确检索到产品名和企业名。",
    },
    "2-4": {
        "reg_no": "国械注准20243020427",
        "reg_name": "一次性腔镜用直线型切割吻合器及组件",
        "class": "III类",
        "holder": "常州威克医疗器械有限公司",
        "scope": "用于腔镜下组织离断、切除和缝合的手动吻合器系统。",
        "status": "已在 NMPA 精确检索到产品名和企业名；具体钉仓组件款式需看附件规格表。",
    },
    "3-1": {
        "reg_no": "待 NMPA 官网按型号进一步核验",
        "reg_name": "待核验（迈瑞 4K三维内窥镜荧光摄影系统）",
        "class": "待核验",
        "holder": "深圳迈瑞生物医疗电子股份有限公司（厂家）",
        "scope": "4K三维内窥镜荧光摄影系统，用于内窥镜手术成像。",
        "status": "已锁定到厂家官网型号 UX5，但未在本轮公开检索中一对一锁定证号。",
    },
    "4-1": {
        "reg_no": "待 NMPA 官网按型号进一步核验",
        "reg_name": "待核验（OMS3500 手术显微镜）",
        "class": "待核验",
        "holder": "苏州速迈医学科技股份有限公司（厂家）",
        "scope": "外科手术显微观察、成像与放大。",
        "status": "已锁定到厂家官网型号 OMS3500，但未在本轮公开检索中一对一锁定证号。",
    },
    "4-2": {
        "reg_no": "待 NMPA 官网按型号进一步核验",
        "reg_name": "待核验（OMS2350 手术显微镜）",
        "class": "待核验",
        "holder": "苏州速迈医学科技股份有限公司（厂家）",
        "scope": "外科手术显微观察、成像与放大。",
        "status": "已锁定到厂家官网型号 OMS2350，但未在本轮公开检索中一对一锁定证号。",
    },
    "5-1": {
        "reg_no": "国械注准20233161058",
        "reg_name": "眼科超声乳化治疗仪",
        "class": "III类",
        "holder": "以诺康医疗科技（苏州）有限公司",
        "scope": "用于白内障等眼前节相关超声乳化治疗场景。",
        "status": "已由 NMPA 公开记录和厂家官网产品线双重核对。",
    },
    "6-1": {
        "reg_no": "待 NMPA 官网按型号进一步核验",
        "reg_name": "待核验（Consona N9 彩色多普勒超声系统）",
        "class": "待核验",
        "holder": "深圳迈瑞生物医疗电子股份有限公司（厂家）",
        "scope": "通科与基础医疗场景彩色多普勒超声检查。",
        "status": "已确认迈瑞存在多张彩超注册证，但未在本轮公开检索中一对一锁定到 Consona N9 对应证号。",
    },
    "6-2": {
        "reg_no": "待 NMPA 官网按型号进一步核验",
        "reg_name": "待核验（昆仑 Resona I9 彩色多普勒超声系统）",
        "class": "待核验",
        "holder": "深圳迈瑞生物医疗电子股份有限公司（厂家）",
        "scope": "腹部、妇产、心脏、血管、介入等多科室彩超检查。",
        "status": "已确认迈瑞存在多张彩超注册证，但未在本轮公开检索中一对一锁定到 Resona I9 对应证号。",
    },
    "7-1": {
        "reg_no": "国械注准20223010108",
        "reg_name": "胸腹腔内窥镜手术系统",
        "class": "III类",
        "holder": "上海微创医疗机器人（集团）股份有限公司",
        "scope": "用于胸腹腔内窥镜手术场景的机器人辅助手术系统。",
        "status": "已由厂家官网和 NMPA 公开记录双重核对；NMPA 近期送达信息仍可检索到该证号。",
    },
    "7-2": {
        "reg_no": "国械注准20223010509",
        "reg_name": "膝关节置换手术导航定位系统",
        "class": "III类",
        "holder": "苏州微创畅行机器人有限公司",
        "scope": "用于膝关节置换术的导航定位与机器人辅助手术。",
        "status": "已由厂家官网和 NMPA 公开记录双重核对；SkyWalker 为商业名，NMPA 注册名为膝关节置换手术导航定位系统。",
    },
    "7-3": {
        "reg_no": "截至 2026-06-18 未直接锁定",
        "reg_name": "待核验（R-ONE / 冠脉介入机器人）",
        "class": "待核验",
        "holder": "待核验",
        "scope": "冠脉介入手术控制/机器人辅助操作。",
        "status": "本轮未在 NMPA 公开检索中直接锁定 R-ONE 或上海微创对应国内注册证；检索到的同类已注册产品主要为 Corindus、Robocath 等，建议后续按具体注册主体再核验。",
    },
    "8-1": {
        "reg_no": "国械注准20153010492",
        "reg_name": "神经外科手术导航系统",
        "class": "III类",
        "holder": "上海复旦数字医疗科技股份有限公司",
        "scope": "用于神经外科手术导航定位、多模态图像融合等场景。",
        "status": "已在 NMPA 精确检索到企业和注册证；excelim-116 与该注册证的逐一型号映射建议再看证书附件或厂家资料。",
    },
    "8-2": {
        "reg_no": "国械注准20153010492",
        "reg_name": "神经外科手术导航系统",
        "class": "III类",
        "holder": "上海复旦数字医疗科技股份有限公司",
        "scope": "用于神经外科手术导航定位、多模态图像融合等场景。",
        "status": "已在 NMPA 精确检索到企业和注册证；NeuroNav-118 与该注册证的逐一型号映射建议再看证书附件或厂家资料。",
    },
}


def main() -> None:
    desktop = Path.home() / "Desktop"
    source = desktop / "agent_products_0617_enriched_detailed.xlsx"
    output = desktop / "agent_products_0617_enriched_regulatory.xlsx"

    wb = load_workbook(source)
    ws = wb["Enriched"]

    start_col = 14
    new_headers = [
        "注册证号 / NMPA",
        "注册证名称",
        "管理类别",
        "注册人 / 持证人",
        "适用范围 / 预期用途",
        "NMPA核验状态",
    ]

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for offset, title in enumerate(new_headers):
        cell = ws.cell(row=2, column=start_col + offset, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=1, column=start_col + offset).alignment = Alignment(vertical="center")

    for row in range(3, ws.max_row + 1):
        serial = ws.cell(row=row, column=1).value
        info = REG_INFO.get(serial)
        if not info:
            continue
        values = [
            info["reg_no"],
            info["reg_name"],
            info["class"],
            info["holder"],
            info["scope"],
            info["status"],
        ]
        for idx, value in enumerate(values, start=start_col):
            cell = ws.cell(row=row, column=idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = {
        "N": 26,
        "O": 28,
        "P": 14,
        "Q": 26,
        "R": 34,
        "S": 40,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    if "Reg_Summary" in wb.sheetnames:
        del wb["Reg_Summary"]
    summary = wb.create_sheet("Reg_Summary")
    summary.append(["拆分序号", "具体型号/产品名", "注册证号 / NMPA", "注册证名称", "NMPA核验状态"])
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in range(3, ws.max_row + 1):
        summary.append(
            [
                ws.cell(row=row, column=1).value,
                ws.cell(row=row, column=8).value,
                ws.cell(row=row, column=14).value,
                ws.cell(row=row, column=15).value,
                ws.cell(row=row, column=19).value,
            ]
        )

    for col, width in {"A": 10, "B": 34, "C": 28, "D": 28, "E": 42}.items():
        summary.column_dimensions[col].width = width
    for row in summary.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(output)
    print(output)


if __name__ == "__main__":
    main()
