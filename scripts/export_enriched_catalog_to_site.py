from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl


WORKBOOK = Path(r"C:\Users\Michael\Desktop\agent_products_0617_enriched_official_verified_links.xlsx")
OUTPUT = Path(r"C:\Users\Michael\Documents\Healthcare\src\lib\site-catalog.generated.ts")
JSON_OUTPUT = Path(r"C:\Users\Michael\Documents\Healthcare\data\site-catalog.generated.json")


CATEGORY_MAP = {
    "超声刀主机耗材": {
        "slug": "ultrasonic-surgery-systems-consumables",
        "nameZh": "超声刀及配套耗材",
        "nameEn": "Ultrasonic Surgery Systems & Consumables",
        "descriptionZh": "超声刀主机、换能器及各类超声刀头产品。",
        "descriptionEn": "Ultrasonic surgery generators, transducers, and compatible disposable handpieces.",
    },
    "一次性切割吻合器": {
        "slug": "surgical-staplers",
        "nameZh": "切割吻合器",
        "nameEn": "Surgical Staplers",
        "descriptionZh": "腹腔镜与开放手术相关的一次性切割吻合器及组件。",
        "descriptionEn": "Disposable stapling and cutting systems for laparoscopic and open surgery workflows.",
    },
    "3D高清腹腔镜": {
        "slug": "endoscopy-imaging-systems",
        "nameZh": "内窥镜成像系统",
        "nameEn": "Endoscopy Imaging Systems",
        "descriptionZh": "腹腔镜手术用 3D / 4K 内窥镜成像与荧光系统。",
        "descriptionEn": "3D, 4K, and fluorescence-enabled endoscopy imaging systems.",
    },
    "显微镜": {
        "slug": "surgical-microscopes",
        "nameZh": "手术显微镜",
        "nameEn": "Surgical Microscopes",
        "descriptionZh": "用于显微外科、五官科和神经外科等场景的手术显微镜。",
        "descriptionEn": "Surgical microscopes for microsurgery, ENT, neurosurgery, and related specialties.",
    },
    "超声乳化仪器": {
        "slug": "ophthalmic-phaco-systems",
        "nameZh": "眼科超乳系统",
        "nameEn": "Ophthalmic Phaco Systems",
        "descriptionZh": "眼科超声乳化与白内障手术相关设备。",
        "descriptionEn": "Ophthalmic phacoemulsification systems for cataract and anterior segment surgery.",
    },
    "高清B超": {
        "slug": "ultrasound-diagnostic-systems",
        "nameZh": "超声诊断设备",
        "nameEn": "Ultrasound Diagnostic Systems",
        "descriptionZh": "通科与高端全身彩色多普勒超声系统。",
        "descriptionEn": "General imaging and premium whole-body diagnostic ultrasound systems.",
    },
    "手术机器人": {
        "slug": "surgical-robotics",
        "nameZh": "手术机器人",
        "nameEn": "Surgical Robotics",
        "descriptionZh": "腹腔镜、骨科与血管介入相关手术机器人平台。",
        "descriptionEn": "Robotic systems for laparoscopic, orthopedic, and vascular interventional surgery.",
    },
    "神经外科导航设备": {
        "slug": "neurosurgical-navigation-systems",
        "nameZh": "神经外科导航设备",
        "nameEn": "Neurosurgical Navigation Systems",
        "descriptionZh": "用于神经外科导航定位和多模态图像融合的设备。",
        "descriptionEn": "Systems for neurosurgical navigation, planning, and multimodal image guidance.",
    },
}


MANUFACTURER_EN = {
    "苏州以诺康": "Innolcon Suzhou",
    "常州威克医疗器械有限公司": "Victor Medical",
    "迈瑞医疗": "Mindray",
    "苏州速迈": "Zumax",
    "微创机器人": "MedBot",
    "上海复旦数字医疗": "Fudan Digital Medical",
}


PRODUCT_OVERRIDES = {
    "1_8": {
        "summaryZh": "官网公开给出 HIFCBSGPL22/35/45 子型号，HIFCBSGPL35 为其中 35cm 规格。",
    },
    "1_9": {
        "summaryZh": "官网公开给出 HIFCBSGPL22/35/45 子型号，HIFCBSGPL45 为其中 45cm 规格。",
    },
    "1_11": {
        "summaryZh": "官网公开给出 IFCBSGPL22/35/45 子型号，IFCBSGPL35 为其中 35cm 规格。",
    },
    "1_12": {
        "summaryZh": "官网公开给出 IFCBSGPL22/35/45 子型号，IFCBSGPL45 为其中 45cm 规格。",
    },
    "5_1": {
        "imageUrl": "https://www.innolcon.com/pic/202308201692501384978.png",
        "sourceUrl": "https://www.innolcon.com/news/html/?30.html%2F=",
        "summaryZh": "ROSEWOOD 眼科超声乳化治疗仪已在以诺康官方新闻稿与注册获批信息中明确出现。",
        "applicationZh": "用于白内障超声乳化摘除及前节相关眼科手术。",
        "specificationsZh": "官网公开名称为 ROSEWOOD 眼科超声乳化治疗仪；官方介绍强调前房稳定性、负压快速监测反馈、稳定高效的流体系统，并提供一次性/可重复使用积液盒及多种型号手术针头。",
    },
    "7_1": {
        "summaryZh": "图迈 Toumai 已由微创机器人官网产品页与 NMPA 注册信息双重对应。",
    },
    "7_2": {
        "imageUrl": "/uploads/images/skywalker-official-crop.jpg",
        "summaryZh": "鸿鹄 SkyWalker 已由微创机器人官网产品页与 NMPA 注册信息双重对应。",
    },
    "7_3": {
        "imageUrl": "/uploads/images/rone-official-crop.jpg",
        "summaryZh": "R-ONE 型号、商业化获批时间与 NMPA 公开送达信息可相互印证。",
    },
    "8_1": {
        "imageUrl": "/uploads/images/excelim-116-embedded.png",
        "sourceUrl": "https://www.digi-medical.com.cn/News/show?id=79&status=1",
        "summaryZh": "excelim-116 为复旦数字医疗公开展示的旗舰型神经外科手术导航系统。",
        "applicationZh": "用于神经外科手术导航定位，支持复杂颅脑手术场景。",
        "specificationsZh": "官网公开型号为 excelim-116；官方新闻介绍其为全新面市旗舰机型，结合多项自主知识产权技术，面向神经外科术中导航应用。",
    },
    "8_2": {
        "imageUrl": "/uploads/images/neuronav-118-embedded.png",
        "sourceUrl": "https://www.digi-medical.com.cn/News/show?id=65&status=1",
        "summaryZh": "NeuroNav-118 已在复旦数字医疗官网新闻与产品页中公开对应。",
        "applicationZh": "用于神经外科手术导航，适用于脑部穿刺、颅内减压等临床场景。",
        "specificationsZh": "官网公开型号为 NeuroNav-118；官方新闻介绍其搭载新版导航软件系统，支持多模态图像融合、血管造影影像导航，并提供标记点注册与探针划线注册方式。",
    },
}


def clean(value: object) -> str:
    return str(value or "").strip()


def split_urls(value: str) -> list[str]:
    return [item.strip() for item in clean(value).splitlines() if item.strip()]


def first_url(value: str) -> str:
    urls = split_urls(value)
    return urls[0] if urls else ""


def is_direct_image(url: str) -> bool:
    lowered = url.lower()
    return lowered.startswith("http") and any(
        lowered.endswith(ext) or f"{ext}?" in lowered
        for ext in [".png", ".jpg", ".jpeg", ".webp", ".gif", ".svg"]
    )


def slugify(seq: str, model: str) -> str:
    parts = re.findall(r"[a-z0-9]+", f"{seq}-{model}".lower())
    return "-".join(parts)[:96] or seq.replace("-", "p")


def placeholder_en(value: str) -> str:
    text = clean(value)
    return f"EN pending: {text}" if text else "EN pending"


def strip_confidence_prefix(text: str) -> str:
    return re.sub(r"^(高|中高|中|较高|较低|低)[。．、；:\s]*", "", clean(text))


def build_summary(row: dict[str, str], model: str) -> str:
    verified_note = strip_confidence_prefix(row["确定性说明"])
    if verified_note:
        return verified_note
    return f"{model} 已在厂家公开资料中出现。"


def build_packaging(row: dict[str, str]) -> str:
    bits = [
        f"价格参考：{clean(row['价格参考'])}" if clean(row["价格参考"]) else "",
        f"注册证号 / NMPA：{clean(row['注册证号 / NMPA'])}" if clean(row["注册证号 / NMPA"]) else "",
        f"注册证名称：{clean(row['注册证名称'])}" if clean(row["注册证名称"]) else "",
        f"管理类别：{clean(row['管理类别'])}" if clean(row["管理类别"]) else "",
        f"注册人 / 持证人：{clean(row['注册人 / 持证人'])}" if clean(row["注册人 / 持证人"]) else "",
        f"NMPA核验状态：{clean(row['NMPA核验状态'])}" if clean(row["NMPA核验状态"]) else "",
    ]
    return "\n".join([bit for bit in bits if bit])


def js(value: object) -> str:
    return json.dumps(value, ensure_ascii=False)


def main() -> None:
    wb = openpyxl.load_workbook(WORKBOOK)
    ws = wb["Enriched"]
    headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]

    categories = []
    seen_categories = set()
    products = []

    for row_index in range(3, ws.max_row + 1):
        values = [ws.cell(row_index, c).value for c in range(1, ws.max_column + 1)]
        row = {headers[i]: clean(values[i]) for i in range(len(headers))}
        source_name = row["原设备/耗材名称"]
        if not source_name:
            continue

        category = CATEGORY_MAP[source_name]
        if category["slug"] not in seen_categories:
            seen_categories.add(category["slug"])
            categories.append(
                {
                    "slug": category["slug"],
                    "nameZh": category["nameZh"],
                    "nameEn": category["nameEn"],
                    "descriptionZh": category["descriptionZh"],
                    "descriptionEn": category["descriptionEn"],
                    "sortOrder": len(categories) + 1,
                }
            )

        seq = row["拆分序号"]
        product_key = seq.replace("-", "_")
        model = row["具体型号/产品名"]
        source_link = first_url(row["来源链接"])
        image_url = row["图片链接"] if is_direct_image(row["图片链接"]) else ""
        overrides = PRODUCT_OVERRIDES.get(product_key, {})

        source_link = overrides.get("sourceUrl", source_link)
        image_url = overrides.get("imageUrl", image_url)
        summary_zh = overrides.get("summaryZh", build_summary(row, model))
        application_zh = overrides.get("applicationZh", row["适用范围 / 预期用途"] or row["使用科室"])
        specifications_zh = overrides.get("specificationsZh", row["关键参数/公开规格"])
        packaging_zh = build_packaging(row)

        products.append(
            {
                "key": product_key,
                "categorySlug": category["slug"],
                "slug": slugify(seq, model),
                "manufacturerZh": row["原厂家"],
                "manufacturerEn": MANUFACTURER_EN.get(row["原厂家"], placeholder_en(row["原厂家"])),
                "model": model,
                "nameZh": model,
                "nameEn": placeholder_en(model),
                "summaryZh": summary_zh,
                "summaryEn": placeholder_en(summary_zh),
                "applicationZh": application_zh,
                "applicationEn": placeholder_en(application_zh),
                "specificationsZh": specifications_zh,
                "specificationsEn": placeholder_en(specifications_zh),
                "packagingZh": packaging_zh,
                "packagingEn": placeholder_en(packaging_zh),
                "imageUrl": image_url,
                "sourceUrl": source_link,
                "featured": 1 if seq.endswith("-1") or seq in {"7-2", "7-3"} else 0,
                "seoTitleZh": model,
                "seoTitleEn": placeholder_en(model),
                "seoDescriptionZh": summary_zh[:140] or model,
                "seoDescriptionEn": placeholder_en(summary_zh[:140] or model),
                "documentAccess": "public",
            }
        )

    lines = [
        "// This file is auto-generated from the approved Enriched workbook.",
        "export const realCategories = " + js(categories) + " as const;",
        "",
        "export const realProducts = " + js(products) + " as const;",
        "",
    ]
    OUTPUT.write_text("\n".join(lines), encoding="utf-8")
    JSON_OUTPUT.write_text(
        json.dumps({"categories": categories, "products": products}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(str(OUTPUT))


if __name__ == "__main__":
    main()
