from pathlib import Path
from urllib.parse import urlparse
from urllib.request import Request, urlopen
import os
import ssl

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill


ROWS = [
    {
        "serial": "1-1",
        "category": "超声刀主机耗材",
        "subtype": "主机/发生器",
        "maker": "苏州以诺康",
        "benchmark": "美国强生",
        "dept": "普外科、胸外科、泌尿外科、妇科",
        "price": "主机约2万元/台（原表）",
        "model": "Y16 超声软组织切割止血手术设备",
        "spec": "官方公开为 Y16 系统主机；与 HP401/HP501 换能器配套使用，为超声刀手柄提供激励信号和高频交流电源。",
        "image_url": "https://www.innolcon.com/pic/202210131665642108290.png",
        "source_url": "https://www.innolcon.com/\nhttps://www.innolcon.com/page/handle01/",
        "confidence": "高。型号 Y16 可在以诺康官网公开页面直接对应。",
    },
    {
        "serial": "1-2",
        "category": "超声刀主机耗材",
        "subtype": "换能器",
        "maker": "苏州以诺康",
        "benchmark": "美国强生",
        "dept": "普外科、胸外科、泌尿外科、妇科",
        "price": "配套部件",
        "model": "HP401 换能器",
        "spec": "适配 Y16/部分 Y20 系统；铝合金外壳、真金镀层电极组件、压电陶瓷定制，官网公开可用于枪式超声刀头。",
        "image_url": "https://www.innolcon.com/pic/202210111665480277245.png",
        "source_url": "https://www.innolcon.com/page/handle01/\nhttps://www.innolcon.com/page/handle02/",
        "confidence": "高。型号 HP401 在官网公开页面明确出现。",
    },
    {
        "serial": "1-3",
        "category": "超声刀主机耗材",
        "subtype": "换能器",
        "maker": "苏州以诺康",
        "benchmark": "美国强生",
        "dept": "普外科、胸外科、泌尿外科、妇科",
        "price": "配套部件",
        "model": "HP501 换能器",
        "spec": "适配 Y16 系统剪式刀头；官网公开提到 HP501 重仅 60g，便于更灵巧精细操作。",
        "image_url": "https://www.innolcon.com/pic/202210111665480277245.png",
        "source_url": "https://www.innolcon.com/page/handle01/",
        "confidence": "高。型号 HP501 在官网公开页面明确出现。",
    },
    {
        "serial": "1-4",
        "category": "超声刀主机耗材",
        "subtype": "一次性刀头",
        "maker": "苏州以诺康",
        "benchmark": "美国强生",
        "dept": "普外科、胸外科、泌尿外科、妇科",
        "price": "一次性耗材约1200元/套（原表）",
        "model": "SG系列 枪式超声刀头",
        "spec": "有效长度 13/22/35/45cm；直径 5mm；官网公开说明可安全凝闭 5mm 血管。",
        "image_url": "https://www.innolcon.com/pic/202210121665558147126.png",
        "source_url": "https://www.innolcon.com/page/knife01/",
        "confidence": "高。SG 系列、长度和 5mm 闭合能力均为官网公开信息。",
    },
    {
        "serial": "1-5",
        "category": "超声刀主机耗材",
        "subtype": "一次性刀头",
        "maker": "苏州以诺康",
        "benchmark": "美国强生",
        "dept": "普外科、胸外科、泌尿外科、妇科",
        "price": "一次性耗材",
        "model": "SS系列 剪式超声刀头",
        "spec": "有效长度 9/17cm；官网公开说明闭合长度达 16.2mm，具备 IAT 自适应反馈技术。",
        "image_url": "https://www.innolcon.com/pic/202210181666085216149.png",
        "source_url": "https://www.innolcon.com/page/jknife01/",
        "confidence": "高。SS 系列及公开参数来自官网页面。",
    },
    {
        "serial": "1-6",
        "category": "超声刀主机耗材",
        "subtype": "一次性刀头",
        "maker": "苏州以诺康",
        "benchmark": "美国强生",
        "dept": "普外科、胸外科、泌尿外科、妇科",
        "price": "一次性耗材",
        "model": "OCBSGBL系列 第二代枪式超声刀头",
        "spec": "有效长度 22/35/45cm；直径 5-7mm；官网公开说明可安全凝闭最大直径 7mm 血管。",
        "image_url": "https://www.innolcon.com/pic/202502241740381419922.png",
        "source_url": "https://www.innolcon.com/page/knife02/",
        "confidence": "高。系列名和关键参数都在官网公开页。",
    },
    {
        "serial": "1-7",
        "category": "超声刀主机耗材",
        "subtype": "7mm集成式刀头",
        "maker": "苏州以诺康",
        "benchmark": "美国强生",
        "dept": "普外科、胸外科、泌尿外科、妇科",
        "price": "一次性耗材",
        "model": "HIFCBSGPL22",
        "spec": "22cm；官网公开说明搭载 Y16 主机，可安全凝闭 7mm 血管；换能器、刀头集成一体化。",
        "image_url": "https://www.innolcon.com/pic/202601301769756351601.jpg",
        "source_url": "https://www.innolcon.com/page/product10/",
        "confidence": "高。官网明确给出 HIFCBSGPL22/35/45。",
    },
    {
        "serial": "1-8",
        "category": "超声刀主机耗材",
        "subtype": "7mm集成式刀头",
        "maker": "苏州以诺康",
        "benchmark": "美国强生",
        "dept": "普外科、胸外科、泌尿外科、妇科",
        "price": "一次性耗材",
        "model": "HIFCBSGPL35",
        "spec": "35cm；官网公开说明搭载 Y16 主机，可安全凝闭 7mm 血管；换能器、刀头集成一体化。",
        "image_url": "https://www.innolcon.com/pic/202601301769756351601.jpg",
        "source_url": "https://www.innolcon.com/page/product10/",
        "confidence": "高。",
    },
    {
        "serial": "1-9",
        "category": "超声刀主机耗材",
        "subtype": "7mm集成式刀头",
        "maker": "苏州以诺康",
        "benchmark": "美国强生",
        "dept": "普外科、胸外科、泌尿外科、妇科",
        "price": "一次性耗材",
        "model": "HIFCBSGPL45",
        "spec": "45cm；官网公开说明搭载 Y16 主机，可安全凝闭 7mm 血管；换能器、刀头集成一体化。",
        "image_url": "https://www.innolcon.com/pic/202601301769756351601.jpg",
        "source_url": "https://www.innolcon.com/page/product10/",
        "confidence": "高。",
    },
    {
        "serial": "1-10",
        "category": "超声刀主机耗材",
        "subtype": "7mm集成式刀头",
        "maker": "苏州以诺康",
        "benchmark": "美国强生",
        "dept": "普外科、胸外科、泌尿外科、妇科",
        "price": "一次性耗材",
        "model": "IFCBSGPL22",
        "spec": "22cm；官网公开说明搭载 Y20 主机，可安全凝闭 7mm 血管；换能器、刀头集成一体化。",
        "image_url": "https://www.innolcon.com/pic/202601301769756351601.jpg",
        "source_url": "https://www.innolcon.com/page/product10/",
        "confidence": "高。官网明确给出 IFCBSGPL22/35/45。",
    },
    {
        "serial": "1-11",
        "category": "超声刀主机耗材",
        "subtype": "7mm集成式刀头",
        "maker": "苏州以诺康",
        "benchmark": "美国强生",
        "dept": "普外科、胸外科、泌尿外科、妇科",
        "price": "一次性耗材",
        "model": "IFCBSGPL35",
        "spec": "35cm；官网公开说明搭载 Y20 主机，可安全凝闭 7mm 血管；换能器、刀头集成一体化。",
        "image_url": "https://www.innolcon.com/pic/202601301769756351601.jpg",
        "source_url": "https://www.innolcon.com/page/product10/",
        "confidence": "高。",
    },
    {
        "serial": "1-12",
        "category": "超声刀主机耗材",
        "subtype": "7mm集成式刀头",
        "maker": "苏州以诺康",
        "benchmark": "美国强生",
        "dept": "普外科、胸外科、泌尿外科、妇科",
        "price": "一次性耗材",
        "model": "IFCBSGPL45",
        "spec": "45cm；官网公开说明搭载 Y20 主机，可安全凝闭 7mm 血管；换能器、刀头集成一体化。",
        "image_url": "https://www.innolcon.com/pic/202601301769756351601.jpg",
        "source_url": "https://www.innolcon.com/page/product10/",
        "confidence": "高。",
    },
    {
        "serial": "2-1",
        "category": "一次性切割吻合器",
        "subtype": "全电动腔镜吻合器",
        "maker": "常州威克医疗器械有限公司",
        "benchmark": "美国强生",
        "dept": "普外科、胸外科、泌尿外科",
        "price": "电动款约3100-4000元/套（原表）",
        "model": "一次性腔镜用直线型全电动切割吻合器及钉仓组件",
        "spec": "官网公开要点：压力实时反馈、关节头长度优化、全流程按键自动操控、击发手柄适配全型号钉仓。",
        "image_url": "https://www.victormedic.com/upload/cn/image/2023-08/col13/1691630678972.png",
        "source_url": "https://www.victormedic.com/col13/9",
        "confidence": "高。产品名可直接落到官网单页，但官网未公开更细的注册型号代码。",
    },
    {
        "serial": "2-2",
        "category": "一次性切割吻合器",
        "subtype": "电动腔镜吻合器",
        "maker": "常州威克医疗器械有限公司",
        "benchmark": "美国强生",
        "dept": "普外科、胸外科、泌尿外科",
        "price": "电动款约3100-4000元/套（原表）",
        "model": "一次性腔镜用直线型电动切割吻合器及钉仓组件",
        "spec": "官网公开要点：ECC 超薄组织专用款，成型高度 1.0mm；三款器械配套三个系列七种不同高度钉仓；具备鹰嘴设计。",
        "image_url": "https://www.victormedic.com/upload/cn/image/2023-08/col13/1691560401655.png",
        "source_url": "https://www.victormedic.com/col13/1",
        "confidence": "高。产品名和关键公开参数都可直接对应官网页面。",
    },
    {
        "serial": "2-3",
        "category": "一次性切割吻合器",
        "subtype": "手动腔镜吻合器（钉匣款）",
        "maker": "常州威克医疗器械有限公司",
        "benchmark": "美国强生",
        "dept": "普外科、胸外科、泌尿外科",
        "price": "手动款约1200-1800元/套（原表）",
        "model": "一次性腔镜用直线型切割吻合器及组件（钉匣款）",
        "spec": "官网公开要点：六种缝钉高度、紧急回刀设计、单手自适应角度调节、三点组织间隙控制机制。",
        "image_url": "https://www.victormedic.com/upload/cn/image/2023-08/col22/1691631598975.png",
        "source_url": "https://www.victormedic.com/col22/105",
        "confidence": "高。可直接落到威克官网单页。",
    },
    {
        "serial": "2-4",
        "category": "一次性切割吻合器",
        "subtype": "手动腔镜吻合器（钉仓组件款）",
        "maker": "常州威克医疗器械有限公司",
        "benchmark": "美国强生",
        "dept": "普外科、胸外科、泌尿外科",
        "price": "手动款约1200-1800元/套（原表）",
        "model": "一次性腔镜用直线型切割吻合器及组件（钉仓组件款）",
        "spec": "公开特征：三排不等高钛钉、水滴型导向钉槽、防滑凸点钉仓表面、鹰嘴设计；与原表卖点高度一致。",
        "image_url": "https://www.victormedic.com/upload/cn/image/2023-07/col13/1689998121185.png",
        "source_url": "https://www.victormedic.com/col22/2\nhttps://www.victormedic.com/col13/list",
        "confidence": "中高。官网列表和特征页能对应到该产品，但公开页面未完整展示注册型号代码。",
    },
    {
        "serial": "3-1",
        "category": "3D高清腹腔镜",
        "subtype": "整机系统",
        "maker": "迈瑞医疗",
        "benchmark": "美国史赛克 Stryker",
        "dept": "普外科、胸外科、妇科、泌尿外科、骨科",
        "price": "全套约64-71万元（原表）",
        "model": "HyPixel UX5 4K三维内窥镜荧光摄影系统",
        "spec": "官方公开为 4K+3D 一体化平台，支持 NIR/荧光成像；新闻稿公开提到 5.6mm 大瞳距双目3D、双路真4K、白光+ICG 荧光一体化、自动旋转回正。",
        "image_url": "https://www.mindray.com/content/dam/xpace/en/products-solutions/products/laparoscopic-products/endoscope-camera-system/ux5/master1.jpg",
        "source_url": "https://www.mindray.com/cn/products/laparoscopic-products/hypixel-ux5\nhttps://www.mindray.com/cn/media-center/news/mindray-releases-a-4k-3d-electronic-thoracoabdominal-endoscope-system",
        "confidence": "高。",
    },
    {
        "serial": "4-1",
        "category": "显微镜",
        "subtype": "高端外科显微镜",
        "maker": "苏州速迈",
        "benchmark": "德国蔡司",
        "dept": "神经外科、脊柱骨科、耳鼻喉、手足外科、口腔显微手术",
        "price": "约35-100万元区间中的高配机型",
        "model": "OMS3500 手术显微镜",
        "spec": "电动连续变倍变焦；内置 4K 影像系统；VarioDist 电动变焦物镜 F=200-600mm；多关节自由悬浮电磁锁控制。",
        "image_url": "https://www.zumax.cn/wp-content/uploads/2024/02/1-7.jpg",
        "source_url": "https://www.zumax.cn/product/oms3500%E6%89%8B%E6%9C%AF%E6%98%BE%E5%BE%AE%E9%95%9C",
        "confidence": "高。",
    },
    {
        "serial": "4-2",
        "category": "显微镜",
        "subtype": "外科手术显微镜",
        "maker": "苏州速迈",
        "benchmark": "德国蔡司",
        "dept": "耳鼻喉、手外科、整形、显微外科",
        "price": "约35-100万元区间中的中高配机型",
        "model": "OMS2350 手术显微镜",
        "spec": "0°-190° 变角双目镜筒；VarioDist 大变焦物镜 190-300mm；高眼点广角目镜；适合五官科与显微外科场景。",
        "image_url": "https://www.zumax.cn/wp-content/uploads/2021/01/%E5%BE%AE%E4%BF%A1%E5%9B%BE%E7%89%87_20210201105943.png",
        "source_url": "https://www.zumax.cn/product/oms2350%E4%BA%94%E5%AE%98%E7%A7%91%E6%98%BE%E5%BE%AE%E9%95%9C",
        "confidence": "高。",
    },
    {
        "serial": "5-1",
        "category": "超声乳化仪器",
        "subtype": "眼科超乳主机",
        "maker": "苏州以诺康",
        "benchmark": "德国蔡司",
        "dept": "眼科、白内障专科",
        "price": "约58-64万元（原表）",
        "model": "ROSEWOOD® 眼科超声乳化手术系统",
        "spec": "官方公开为眼科超声乳化手术系统；公开卖点包括前房稳定、术中精控、液流稳定、操作便捷；2021 年获南德欧盟 CE 认证。",
        "image_url": "https://www.innolcon.com/templates/default/images/icon31.png",
        "source_url": "https://www.innolcon.com/page/ykcpx/\nhttps://www.innolcon.com/news/html/?30.html%2F=",
        "confidence": "高。",
    },
    {
        "serial": "6-1",
        "category": "高清B超",
        "subtype": "中端全科机",
        "maker": "迈瑞医疗",
        "benchmark": "飞利浦",
        "dept": "腹部、妇产、浅表、小器官等通科场景",
        "price": "中端约50万元左右（原表）",
        "model": "Consona N9 彩色多普勒超声系统",
        "spec": "官方为 Consona N 系列机型；公开卖点包括 ZST+ 域光成像平台、面向基础医疗与通科筛查的智慧型彩超定位。",
        "image_url": "https://www.mindray.com/content/dam/xpace/en/products-solutions/products/ultrasound/primary-care/consona-n-series/consona-n9-kv-pc.jpg",
        "source_url": "https://www.mindray.com/cn/products/ultrasound/primary-care/consona-n-series",
        "confidence": "中高。原表未给唯一型号，这里按中端价格带与全科定位落到 N9。",
    },
    {
        "serial": "6-2",
        "category": "高清B超",
        "subtype": "中高端全身机",
        "maker": "迈瑞医疗",
        "benchmark": "飞利浦",
        "dept": "腹部、妇产、心脏、血管、介入等多科室",
        "price": "中高端约72-90万元（原表）",
        "model": "昆仑 Resona I9 彩色多普勒超声系统",
        "spec": "官方为高端全身机；原表所述 AI 调参、弹性成像、三维四维、向量血流等特征与 Resona I9 所在高端通用彩超平台更匹配。",
        "image_url": "https://www.mindray.com/content/dam/xpace/zh/products-solutions/products/ultrasound/general-imaging/resona-i9/zhg997-s1-web.jpg",
        "source_url": "https://www.mindray.com/cn/products/ultrasound/general-imaging/resona-i9",
        "confidence": "中高。是按原表中高端价位段和功能描述落到最匹配型号。",
    },
    {
        "serial": "7-1",
        "category": "手术机器人",
        "subtype": "腔镜手术机器人",
        "maker": "微创机器人",
        "benchmark": "达芬奇",
        "dept": "泌尿外科、妇科、普外科、胸外科",
        "price": "约1100-1500万元（原表）",
        "model": "图迈® Toumai® 胸腹腔内窥镜手术系统",
        "spec": "官方公开注册证号：国械注准 20223010108；由医生控制台、患者手术平台和图像平台组成；支持远程手术网络应用。",
        "image_url": "https://www.medbotsurgical.com/upload/solution/1673475044165299373.png",
        "source_url": "https://www.medbotsurgical.com/solution/1.html\nhttps://www.medbotsurgical.com/solution",
        "confidence": "高。",
    },
    {
        "serial": "7-2",
        "category": "手术机器人",
        "subtype": "骨科手术机器人",
        "maker": "微创机器人",
        "benchmark": "达芬奇/骨科进口机器人",
        "dept": "骨科、关节外科",
        "price": "约550-700万元（原表）",
        "model": "鸿鹄® SkyWalker™ 骨科手术机器人",
        "spec": "官方公开注册证号：国械注准 20223010509；基于 CT 与假体模型做术前规划，术中配准+机械臂辅助截骨，可避免传统髓腔定位。",
        "image_url": "https://www.medbotsurgical.com/upload/solution/1673805991456891055.jpg",
        "source_url": "https://www.medbotsurgical.com/solution/3.html\nhttps://www.medbotsurgical.com/solution",
        "confidence": "高。",
    },
    {
        "serial": "7-3",
        "category": "手术机器人",
        "subtype": "血管介入机器人",
        "maker": "微创机器人",
        "benchmark": "进口介入机器人",
        "dept": "心内科、介入科",
        "price": "约800-1000万元（原表）",
        "model": "R-ONE® 血管介入机器人",
        "spec": "官方公开为 PCI 机器人产品；公开卖点包括双手仿生搓捻、0-360° 旋转平移连续运动、亚毫米级步进、一键锁定解锁。",
        "image_url": "https://www.medbotsurgical.com/upload/news/1702433671135377295.jpg",
        "source_url": "https://www.medbotsurgical.com/news/293.html\nhttps://www.medbotsurgical.com/news/108.html",
        "confidence": "高。",
    },
    {
        "serial": "8-1",
        "category": "神经外科导航设备",
        "subtype": "神经外科手术导航系统",
        "maker": "上海复旦数字医疗",
        "benchmark": "美敦力",
        "dept": "神经外科",
        "price": "约150-200万元（原表）",
        "model": "excelim-116 神经外科手术导航系统",
        "spec": "复旦数字医疗官网公开在产品中心列示；第87届CMEF新闻稿将其列为公司重点展示的神经外科手术导航系统型号之一。",
        "image_url": "https://image.made-in-china.com/202f0j00znQlTqcRsyUD/Surgical-Navigation-System-Neurosurgery-Medical-Instrument-with-Surgical-Set.webp",
        "source_url": "https://www.digi-medical.com.cn/Product/show?id=1\nhttps://www.digi-medical.com.cn/News/show?id=79&status=1",
        "confidence": "中高。型号来自官网产品页；受站点限制，预览图采用公开网络实物图，来源链接保留官网。",
    },
    {
        "serial": "8-2",
        "category": "神经外科导航设备",
        "subtype": "神经外科手术导航系统",
        "maker": "上海复旦数字医疗",
        "benchmark": "美敦力",
        "dept": "神经外科",
        "price": "约150-200万元（原表）",
        "model": "NeuroNav-118 神经外科手术导航系统",
        "spec": "官网公开型号；新闻稿公开提到支持多模态图像融合、血管造影影像导航，并提供探针划线注册等功能。",
        "image_url": "https://image.made-in-china.com/202f0j00rFQlajqyCNUI/Surgical-Navigation-System-Neurosurgery-Medical-Instrument-with-Surgical-Set.webp",
        "source_url": "https://www.digi-medical.com.cn/Product/show?id=3\nhttps://www.digi-medical.com.cn/News/show?id=65&status=1",
        "confidence": "中高。型号和功能点来自官网；预览图采用公开网络实物图，来源链接保留官网。",
    },
]


def download(url: str, dest: Path) -> bool:
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    try:
        with urlopen(req, timeout=30, context=ctx) as response, open(dest, "wb") as fh:
            fh.write(response.read())
        return True
    except Exception:
        return False


def main() -> None:
    desktop = Path.home() / "Desktop"
    source_candidates = [
        p
        for p in desktop.glob("*0617*.xlsx")
        if not p.name.startswith("~$") and not p.name.startswith("agent_products_")
    ]
    source = source_candidates[0]
    output = desktop / "agent_products_0617_enriched_detailed.xlsx"
    img_dir = Path(r"C:\Users\Michael\Documents\Healthcare\_tmp_product_images_detailed")
    img_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(source)
    for sheet_name in ["Enriched", "Sources"]:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    ws = wb.create_sheet("Enriched")
    ws["A1"] = "代理产品信息完善建议（按具体型号/具体产品逐行拆分，基于 2026-06-17 公开资料整理）"

    headers = [
        "拆分序号",
        "原设备/耗材名称",
        "细分类型",
        "原厂家",
        "对标厂家",
        "使用科室",
        "价格参考",
        "具体型号/产品名",
        "关键参数/公开规格",
        "图片链接",
        "来源链接",
        "确定性说明",
        "图片预览",
    ]
    for col_idx, title in enumerate(headers, start=1):
        ws.cell(row=2, column=col_idx, value=title)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    sub_fill = PatternFill("solid", fgColor="D9EAF7")

    for col in range(1, 14):
        top = ws.cell(row=1, column=col)
        top.fill = sub_fill
        top.font = Font(bold=True)
        top.alignment = Alignment(vertical="center")

        header = ws.cell(row=2, column=col)
        header.fill = header_fill
        header.font = header_font
        header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_idx = 3
    for item in ROWS:
        values = [
            item["serial"],
            item["category"],
            item["subtype"],
            item["maker"],
            item["benchmark"],
            item["dept"],
            item["price"],
            item["model"],
            item["spec"],
            item["image_url"],
            item["source_url"],
            item["confidence"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws.cell(row=row_idx, column=10).hyperlink = item["image_url"]
        ws.cell(row=row_idx, column=10).style = "Hyperlink"
        ws.cell(row=row_idx, column=11).hyperlink = item["source_url"].split("\n")[0]
        ws.cell(row=row_idx, column=11).style = "Hyperlink"

        path_part = urlparse(item["image_url"]).path
        ext = os.path.splitext(path_part)[1] or ".png"
        local_img = img_dir / f"{item['serial'].replace('-', '_')}{ext}"
        if ext.lower() in {".png", ".jpg", ".jpeg", ".webp"} and download(item["image_url"], local_img):
            image = XLImage(str(local_img))
            max_width = 170
            scale = max_width / image.width
            image.width = max_width
            image.height = int(image.height * scale)
            ws.add_image(image, f"M{row_idx}")
            ws.row_dimensions[row_idx].height = max(110, image.height * 0.75)
        else:
            ws.cell(row=row_idx, column=13, value="点击图片链接查看")
            ws.cell(row=row_idx, column=13).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row_idx].height = 90

        row_idx += 1

    widths = {
        "A": 10,
        "B": 18,
        "C": 18,
        "D": 16,
        "E": 18,
        "F": 20,
        "G": 18,
        "H": 32,
        "I": 44,
        "J": 30,
        "K": 32,
        "L": 28,
        "M": 26,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:M{ws.max_row}"

    ref = wb.create_sheet("Sources")
    ref.append(["拆分序号", "具体型号/产品名", "来源链接"])
    for cell in ref[1]:
        cell.fill = header_fill
        cell.font = header_font
    for item in ROWS:
        ref.append([item["serial"], item["model"], item["source_url"]])
    for col, width in {"A": 10, "B": 40, "C": 90}.items():
        ref.column_dimensions[col].width = width
    for row in ref.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(output)
    print(output)


if __name__ == "__main__":
    main()
