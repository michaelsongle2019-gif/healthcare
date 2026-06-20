from __future__ import annotations

from copy import copy
from pathlib import Path

import openpyxl


SRC = Path(r"C:\Users\Michael\Desktop\agent_products_0617_enriched_regulatory.xlsx")
OUT = Path(r"C:\Users\Michael\Desktop\agent_products_0617_enriched_official_verified.xlsx")


UPDATES = {
    "3-1": {
        "具体型号/产品名": "HyPixel UX5 4K三维内窥镜荧光摄影系统",
        "关键参数/公开规格": (
            "迈瑞官网与官方新闻页公开要点：5.6mm 大瞳距双目 3D 视觉，双路真 4K 图像链路，"
            "支持白光 + ICG 荧光成像一体化，并具备镜头自动旋转回正。"
        ),
        "来源链接": (
            "https://www.mindray.com/en/products/laparoscopic-products/hypixel-ux5\n"
            "https://www.mindray.com/cn/media-center/news/mindray-releases-a-4k-3d-electronic-thoracoabdominal-endoscope-system"
        ),
        "确定性说明": "高。UX5 型号与 4K/3D/荧光特征可由迈瑞官方英文产品页和官方新闻页交叉确认。",
    },
    "4-1": {
        "关键参数/公开规格": (
            "速迈官网公开参数：电动变倍系统 0.4x-2.4x（1:6），总放大倍数 1.3x-18.5x；"
            "VarioDist 电动变焦物镜工作距离 200-600mm；0°-190° 变角双目镜筒；"
            "200mm 工作距离下照度≥100000 lx；可选内置 4K / 4K-3D 影像系统。"
        ),
        "确定性说明": "高。OMS3500 型号和主要光学参数均可在速迈官网产品页直接核对。",
    },
    "4-2": {
        "关键参数/公开规格": (
            "速迈官网公开参数：0°-190° 变角双目镜筒，高眼点广角目镜；"
            "VarioDist 大变焦物镜 190-300mm；6 档手动变倍；支持精确瞳距调节，定位于五官科与显微外科。"
        ),
        "确定性说明": "高。OMS2350 型号及主要结构参数可在速迈官网产品页直接核对。",
    },
    "5-1": {
        "图片链接": "https://www.innolcon.com/news/html/?30.html%2F=",
        "来源链接": (
            "https://www.innolcon.com/page/ykcpx/\n"
            "https://www.innolcon.com/news/html/?30.html%2F="
        ),
        "确定性说明": "高。ROSEWOOD 商业名、产品性质与官方新闻稿一致；图片链接改为官方新闻页实拍图页面。",
    },
    "6-1": {
        "来源链接": (
            "https://www.mindray.com/en/products/ultrasound/primary-care/consona-n9\n"
            "https://www.mindray.com/etc.clientlibs/xpace/clientlibs/clientlib-site/resources/plugins/web/viewer.html?file=%2Fcontent%2Fdam%2Fxpace%2Fen%2Fresources%2Fbrochure%2Fconsona-n9-product-brochure.pdf\n"
            "https://www.mindray.com/content/dam/xpace/en_us/products-solutions/products/ultrasound-machines/consona-n9/resources/literature/Consona_N9__transducers_40756_Rev1.0.pdf"
        ),
        "关键参数/公开规格": (
            "已切到迈瑞官方单机页面 Consona N9；官方公开定位为 shared service / primary care 场景彩超，"
            "并配有官方 brochure 与探头族文档。原表未给唯一型号，此处按中端全科定位收敛到 N9。"
        ),
        "确定性说明": "中高。Consona N9 为迈瑞官方真实机型，但原表未给唯一型号，本行仍属于按价格带和功能定位后的最佳匹配。",
    },
    "6-2": {
        "来源链接": (
            "https://www.mindray.com/cn/products/ultrasound/general-imaging/resona-i9\n"
            "https://www.mindray.com/content/dam/xpace/en_us/products-solutions/products/ultrasound-machines/images/pdf/Mindray-Resona-I9-brochure-Rev-B_CR46819.pdf"
        ),
        "关键参数/公开规格": (
            "迈瑞官方页面与 brochure 公开要点：Resona I9 为高端全身彩超平台，"
            "具备模块化设计、连续扫描电池、低噪音平台，并对应高端通用超声场景。"
        ),
        "确定性说明": "中高。Resona I9 为迈瑞官方真实机型；但原表未给唯一型号，本行是按价位段与功能描述收敛后的最佳匹配。",
    },
    "7-2": {
        "具体型号/产品名": "鸿鹄® SkyWalker™ 膝关节置换手术导航定位系统",
        "关键参数/公开规格": (
            "微创机器人官网公开注册证号：国械注准20223010509；"
            "基于 CT 与假体模型进行术前规划，术中采用配准 + 轻量化机械臂辅助截骨，"
            "可避免传统髓腔定位，主要用于膝关节置换术。"
        ),
        "来源链接": (
            "https://www.medbotsurgical.com/solution/3.html\n"
            "https://www.medbotsurgical.com/news/17.html\n"
            "https://www.medbotsurgical.com/solution"
        ),
        "适用范围 / 预期用途": "用于膝关节置换术的导航定位与机器人辅助手术。",
        "NMPA核验状态": "已由厂家官网与公开监管送达信息交叉核对；SkyWalker 为商业名，注册名为膝关节置换手术导航定位系统。",
    },
    "7-3": {
        "具体型号/产品名": "R-ONE® 冠状动脉介入手术控制系统",
        "关键参数/公开规格": (
            "微创机器人官方新闻与 NMPA 公开送达信息显示：R-ONE 为冠脉 PCI 机器人产品；"
            "具备双手仿生搓捻、0-360° 连续旋转/平移、亚毫米级步进、一键锁定解锁，"
            "适配主流介入器械和导管室。"
        ),
        "来源链接": (
            "https://www.medbotsurgical.com/en/news/224.html\n"
            "https://www.nmpa.gov.cn/directory/web/nmpa/zwfw/sdxx/sdxxylqx/qxpjfb/20250127095542145.html\n"
            "https://www.nmpa.gov.cn/directory/web/nmpa////zwfw/sdxx/sdxxylqx/qxpjfb/20231212142716117.html"
        ),
        "注册证号 / NMPA": "国械注进20233010579",
        "注册证名称": "冠状动脉介入手术控制系统",
        "管理类别": "III类",
        "注册人 / 持证人": "Robocath S.A.S / 知脉（上海）机器人有限公司",
        "适用范围 / 预期用途": "用于经皮冠状动脉介入手术（PCI）期间，对导丝/导管相关器械进行机器人辅助控制。",
        "NMPA核验状态": "已依据微创机器人官方获批新闻与 NMPA 公开送达信息锁定到进口注册证国械注进20233010579。",
        "确定性说明": "高。R-ONE 型号、商业化获批时间与 NMPA 公开送达信息可相互印证。",
    },
    "8-1": {
        "图片链接": "https://www.digi-medical.com.cn/News/show?id=79&status=1",
        "来源链接": (
            "https://www.digi-medical.com.cn/Product/show?id=1\n"
            "https://www.digi-medical.com.cn/News/show?id=79&status=1\n"
            "https://www.digi-medical.com.cn/News/show?id=76&status=1"
        ),
        "关键参数/公开规格": (
            "官网产品中心可确认 excelim-116 为神经外科手术导航系统；"
            "官方展会新闻与注册相关新闻可确认其为公司重点导航设备型号之一。"
        ),
        "确定性说明": "中高。型号与产品属性来自官网；图片链接改为官方新闻实拍页，因官网站点限制未稳定暴露原图地址。",
    },
    "8-2": {
        "图片链接": "https://www.digi-medical.com.cn/News/show?id=65&status=1",
        "来源链接": (
            "https://www.digi-medical.com.cn/Product/show?id=3\n"
            "https://www.digi-medical.com.cn/News/show?id=65&status=1"
        ),
        "关键参数/公开规格": (
            "官网产品页可确认 NeuroNav-118 为神经外科手术导航系统；"
            "官方新闻公开其适用于脑部穿刺类场景，并已出口土耳其市场。"
        ),
        "确定性说明": "中高。型号与产品属性来自官网；图片链接改为官方新闻实拍页，因官网站点限制未稳定暴露原图地址。",
    },
}


def build_header_map(sheet):
    return {sheet.cell(2, col).value: col for col in range(1, sheet.max_column + 1)}


def set_cell(sheet, row: int, col: int, value):
    target = sheet.cell(row, col)
    target.value = value


def main():
    wb = openpyxl.load_workbook(SRC)
    enriched = wb["Enriched"]
    sources = wb["Sources"]
    reg = wb["Reg_Summary"]

    enriched["A1"] = "代理产品信息完善建议（按具体型号/具体产品逐行拆分，基于 2026-06-18 官方公开资料复核）"

    enriched_map = build_header_map(enriched)
    source_map = {sources.cell(r, 1).value: r for r in range(2, sources.max_row + 1)}
    reg_map = {reg.cell(r, 1).value: r for r in range(2, reg.max_row + 1)}

    for row in range(3, enriched.max_row + 1):
        seq = enriched.cell(row, 1).value
        if seq not in UPDATES:
            continue
        for header, value in UPDATES[seq].items():
            col = enriched_map[header]
            set_cell(enriched, row, col, value)

        if seq in source_map:
            src_row = source_map[seq]
            if "具体型号/产品名" in UPDATES[seq]:
                sources.cell(src_row, 2).value = UPDATES[seq]["具体型号/产品名"]
            if "来源链接" in UPDATES[seq]:
                sources.cell(src_row, 3).value = UPDATES[seq]["来源链接"]

        if seq in reg_map:
            reg_row = reg_map[seq]
            if "具体型号/产品名" in UPDATES[seq]:
                reg.cell(reg_row, 2).value = UPDATES[seq]["具体型号/产品名"]
            if "注册证号 / NMPA" in UPDATES[seq]:
                reg.cell(reg_row, 3).value = UPDATES[seq]["注册证号 / NMPA"]
            if "注册证名称" in UPDATES[seq]:
                reg.cell(reg_row, 4).value = UPDATES[seq]["注册证名称"]
            if "NMPA核验状态" in UPDATES[seq]:
                reg.cell(reg_row, 5).value = UPDATES[seq]["NMPA核验状态"]

    wb.save(OUT)
    print(str(OUT))


if __name__ == "__main__":
    main()
